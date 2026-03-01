"""
Generate viability analysis spreadsheet templates.

Usage:
    python build_spreadsheets.py <output_folder>

Generates:
    - viability-scorecard.xlsx (weighted scorecard with formulas)
    - phase-3-competitive-data.xlsx (competitor comparison + gap analysis + pricing)
"""

import sys
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

if len(sys.argv) < 2:
    print("Usage: python build_spreadsheets.py <output_folder>")
    sys.exit(1)

output_folder = sys.argv[1]
os.makedirs(output_folder, exist_ok=True)

# Shared styles
HEADER_FILL = PatternFill('solid', fgColor='1A1A2E')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
SUBHEADER_FILL = PatternFill('solid', fgColor='16213E')
SUBHEADER_FONT = Font(name='Arial', bold=True, color='E0E0E0', size=10)
INPUT_FILL = PatternFill('solid', fgColor='FFF8E1')
INPUT_FONT = Font(name='Arial', color='0000FF', size=11)
FORMULA_FONT = Font(name='Arial', color='000000', size=11)
BOLD_FONT = Font(name='Arial', bold=True, size=11)
NORMAL_FONT = Font(name='Arial', size=11)
TITLE_FONT = Font(name='Arial', bold=True, size=16, color='1A1A2E')
SUBTITLE_FONT = Font(name='Arial', size=11, color='666666')
GREEN_FILL = PatternFill('solid', fgColor='C8E6C9')
YELLOW_FILL = PatternFill('solid', fgColor='FFF9C4')
ORANGE_FILL = PatternFill('solid', fgColor='FFE0B2')
RED_FILL = PatternFill('solid', fgColor='FFCDD2')
BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)


def build_scorecard(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Scorecard"

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 40

    ws.merge_cells('A1:F1')
    ws['A1'] = 'Pre-Build Viability Scorecard'
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(vertical='center')
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:F2')
    ws['A2'] = 'Project: [Enter project name]'
    ws['A2'].font = SUBTITLE_FONT

    ws.merge_cells('A3:F3')
    ws['A3'] = 'Date: [YYYY-MM-DD]'
    ws['A3'].font = SUBTITLE_FONT

    ws.merge_cells('A5:F5')
    ws['A5'] = 'Scoring: 1 = Red flag  |  2 = Weak  |  3 = Acceptable  |  4 = Good  |  5 = Excellent'
    ws['A5'].font = Font(name='Arial', italic=True, size=10, color='666666')

    headers = ['#', 'Dimension', 'Score (1-5)', 'Weight', 'Weighted Score', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER
    ws.row_dimensions[7].height = 30

    dimensions = [
        (1, 'Problem severity — Is the pain real and frequent?', 3),
        (2, 'Persona clarity — Can you name and find the buyer?', 2),
        (3, 'Market size — Enough potential customers?', 2),
        (4, 'Competitive gap — Is there a real opening?', 3),
        (5, 'Differentiation — Why you, why now?', 2),
        (6, 'Business model — Do the numbers work?', 3),
        (7, 'Acquisition channel — Can you reach customers affordably?', 2),
        (8, 'Technical feasibility — Can you build it with your stack?', 1),
        (9, 'Founder-market fit — Do you have relevant expertise?', 2),
        (10, 'Solo founder viability — Can one person pull this off?', 2),
    ]

    for i, (num, dim, weight) in enumerate(dimensions):
        row = 8 + i
        ws.cell(row=row, column=1, value=num).font = BOLD_FONT
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=1).border = BORDER
        ws.cell(row=row, column=2, value=dim).font = NORMAL_FONT
        ws.cell(row=row, column=2).border = BORDER
        score_cell = ws.cell(row=row, column=3)
        score_cell.fill = INPUT_FILL
        score_cell.font = INPUT_FONT
        score_cell.alignment = Alignment(horizontal='center')
        score_cell.border = BORDER
        ws.cell(row=row, column=4, value=weight).font = FORMULA_FONT
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=4).border = BORDER
        ws.cell(row=row, column=5, value=f'=IF(C{row}="","",C{row}*D{row})').font = FORMULA_FONT
        ws.cell(row=row, column=5).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5).border = BORDER
        ws.cell(row=row, column=6).fill = INPUT_FILL
        ws.cell(row=row, column=6).font = INPUT_FONT
        ws.cell(row=row, column=6).border = BORDER

    total_row = 18
    ws.cell(row=total_row, column=2, value='TOTAL').font = SUBHEADER_FONT
    ws.cell(row=total_row, column=2).fill = SUBHEADER_FILL
    ws.cell(row=total_row, column=2).border = BORDER
    for col in [1, 3, 6]:
        ws.cell(row=total_row, column=col).fill = SUBHEADER_FILL
        ws.cell(row=total_row, column=col).border = BORDER
    ws.cell(row=total_row, column=4, value='=SUM(D8:D17)').font = SUBHEADER_FONT
    ws.cell(row=total_row, column=4).fill = SUBHEADER_FILL
    ws.cell(row=total_row, column=4).alignment = Alignment(horizontal='center')
    ws.cell(row=total_row, column=4).border = BORDER
    ws.cell(row=total_row, column=5, value='=SUM(E8:E17)').font = Font(name='Arial', bold=True, color='FFFFFF', size=12)
    ws.cell(row=total_row, column=5).fill = SUBHEADER_FILL
    ws.cell(row=total_row, column=5).alignment = Alignment(horizontal='center')
    ws.cell(row=total_row, column=5).border = BORDER

    ws.cell(row=19, column=2, value='Max Possible').font = Font(name='Arial', italic=True, size=10, color='999999')
    ws.cell(row=19, column=5, value=110).font = Font(name='Arial', italic=True, size=10, color='999999')
    ws.cell(row=19, column=5).alignment = Alignment(horizontal='center')

    ws.cell(row=20, column=2, value='Percentage').font = BOLD_FONT
    ws.cell(row=20, column=5, value='=IF(E18="","",E18/110)').font = BOLD_FONT
    ws.cell(row=20, column=5).number_format = '0%'
    ws.cell(row=20, column=5).alignment = Alignment(horizontal='center')

    ws.merge_cells('A22:F22')
    ws['A22'] = 'DECISION'
    ws['A22'].font = Font(name='Arial', bold=True, size=14, color='1A1A2E')

    ws.cell(row=23, column=2, value='Auto Decision:').font = BOLD_FONT
    ws.cell(row=23, column=3, value='=IF(E18="","Enter scores above",IF(E18>=88,"🟢 STRONG GO",IF(E18>=66,"🟡 CONDITIONAL GO",IF(E18>=44,"🟠 PIVOT","🔴 KILL"))))').font = Font(name='Arial', bold=True, size=13)
    ws.merge_cells('C23:F23')

    ref_data = [
        ('88-110 (80%+)', '🟢 Strong Go', 'Proceed to SaaS Intake Questionnaire', GREEN_FILL),
        ('66-87 (60-79%)', '🟡 Conditional Go', 'Address weak areas, re-run weak phases', YELLOW_FILL),
        ('44-65 (40-59%)', '🟠 Pivot', 'Rethink positioning, persona, or problem', ORANGE_FILL),
        ('Below 44 (<40%)', '🔴 Kill', 'Move to next idea. Keep research for later', RED_FILL),
    ]
    for h_col, h_val in [(2, 'Score Range'), (3, 'Decision'), (4, 'Action')]:
        ws.cell(row=25, column=h_col, value=h_val).font = HEADER_FONT
        ws.cell(row=25, column=h_col).fill = HEADER_FILL
        ws.cell(row=25, column=h_col).border = BORDER
    ws.merge_cells('D25:F25')

    for i, (sr, dec, act, fill) in enumerate(ref_data):
        row = 26 + i
        for col, val in [(2, sr), (3, dec), (4, act)]:
            c = ws.cell(row=row, column=col, value=val)
            c.font = BOLD_FONT if col == 3 else NORMAL_FONT
            c.fill = fill
            c.border = BORDER
        ws.merge_cells(f'D{row}:F{row}')

    ws.merge_cells('A31:F31')
    ws['A31'] = 'ASSESSMENT'
    ws['A31'].font = Font(name='Arial', bold=True, size=14, color='1A1A2E')
    for row, label in [(32, 'Strongest dimensions:'), (33, 'Weakest dimensions:'), (34, 'Key risks to monitor:'), (35, 'Pivot options:')]:
        ws.cell(row=row, column=2, value=label).font = BOLD_FONT
        ws.merge_cells(f'C{row}:F{row}')
        ws.cell(row=row, column=3).fill = INPUT_FILL
        ws.cell(row=row, column=3).font = INPUT_FONT
        ws.cell(row=row, column=3).border = BORDER

    wb.save(path)


def build_competitive_data(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Competitor Comparison"

    for col, w in {'A': 22, 'B': 18, 'C': 16, 'D': 35, 'E': 35, 'F': 14, 'G': 14, 'H': 14, 'I': 35, 'J': 35}.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells('A1:J1')
    ws['A1'] = 'Competitive Landscape Data — [Project Name]'
    ws['A1'].font = TITLE_FONT
    ws.row_dimensions[1].height = 35
    ws.merge_cells('A2:J2')
    ws['A2'] = 'Date: [YYYY-MM-DD]'
    ws['A2'].font = SUBTITLE_FONT

    comp_headers = ['Competitor', 'Type', 'Pricing', 'Key Features', 'Weaknesses (from reviews)',
                    'Product Quality\n(1-5)', 'Pricing Accessibility\n(1-5)', 'Market Overlap\n(1-5)',
                    'User Complaints', 'Notes']
    for col, header in enumerate(comp_headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[4].height = 40

    for row in range(5, 15):
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.border = BORDER
            cell.font = NORMAL_FONT
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if col in (6, 7, 8):
                cell.fill = INPUT_FILL
                cell.alignment = Alignment(horizontal='center', vertical='top')

    for row, label in [(5, 'Direct'), (6, 'Direct'), (7, 'Direct'), (8, 'Indirect'), (9, 'Indirect'), (10, 'Adjacent')]:
        ws.cell(row=row, column=2, value=label).font = Font(name='Arial', italic=True, color='999999', size=10)

    avg_row = 15
    ws.cell(row=avg_row, column=1, value='AVERAGES').font = SUBHEADER_FONT
    ws.cell(row=avg_row, column=1).fill = SUBHEADER_FILL
    for col in range(2, 11):
        ws.cell(row=avg_row, column=col).fill = SUBHEADER_FILL
        ws.cell(row=avg_row, column=col).border = BORDER
    for col in (6, 7, 8):
        letter = get_column_letter(col)
        cell = ws.cell(row=avg_row, column=col)
        cell.value = f'=IF(COUNTA({letter}5:{letter}14)=0,"",AVERAGE({letter}5:{letter}14))'
        cell.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        cell.number_format = '0.0'
        cell.alignment = Alignment(horizontal='center')

    # Gap Analysis sheet
    ws2 = wb.create_sheet('Gap Analysis')
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 50
    ws2.column_dimensions['C'].width = 20
    ws2.merge_cells('A1:C1')
    ws2['A1'] = 'Market Gap Analysis'
    ws2['A1'].font = TITLE_FONT
    for col, header in enumerate(['Gap Identified', 'Description & Evidence', 'Opportunity Level'], 1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
    for row in range(4, 12):
        for col in range(1, 4):
            cell = ws2.cell(row=row, column=col)
            cell.border = BORDER
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if col == 3:
                cell.fill = INPUT_FILL
                cell.alignment = Alignment(horizontal='center', vertical='top')

    # Pricing Landscape sheet
    ws3 = wb.create_sheet('Pricing Landscape')
    ws3.column_dimensions['A'].width = 22
    ws3.column_dimensions['B'].width = 18
    ws3.column_dimensions['C'].width = 18
    ws3.column_dimensions['D'].width = 18
    ws3.column_dimensions['E'].width = 30
    ws3.merge_cells('A1:E1')
    ws3['A1'] = 'Pricing Landscape'
    ws3['A1'].font = TITLE_FONT
    for col, header in enumerate(['Competitor', 'Free Tier', 'Low Tier (€/mo)', 'High Tier (€/mo)', 'Pricing Model Notes'], 1):
        cell = ws3.cell(row=3, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
    for row in range(4, 14):
        for col in range(1, 6):
            cell = ws3.cell(row=row, column=col)
            cell.border = BORDER
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    avg_row = 14
    ws3.cell(row=avg_row, column=1, value='MARKET AVERAGE').font = SUBHEADER_FONT
    ws3.cell(row=avg_row, column=1).fill = SUBHEADER_FILL
    for col in range(2, 6):
        ws3.cell(row=avg_row, column=col).fill = SUBHEADER_FILL
        ws3.cell(row=avg_row, column=col).border = BORDER
    for col in (3, 4):
        letter = get_column_letter(col)
        cell = ws3.cell(row=avg_row, column=col)
        cell.value = f'=IF(COUNTA({letter}4:{letter}13)=0,"",AVERAGE({letter}4:{letter}13))'
        cell.font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
        cell.number_format = '€#,##0'

    wb.save(path)


scorecard_path = os.path.join(output_folder, 'viability-scorecard.xlsx')
competitive_path = os.path.join(output_folder, 'phase-3-competitive-data.xlsx')

build_scorecard(scorecard_path)
build_competitive_data(competitive_path)

print(f"Created: {scorecard_path}")
print(f"Created: {competitive_path}")
